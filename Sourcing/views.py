from django.shortcuts import render, get_object_or_404
from django.http import HttpResponseRedirect
from django.http import Http404, HttpResponse, JsonResponse
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from .models import *
from RFP.models import *
from State.models import *
from Customer.models import *
from Supplier.models import *
from Employee.models import *
from django.core.mail import send_mail, EmailMessage
from django.core import mail
import random
from django.conf import settings
import openpyxl

from rest_framework.response import Response

from django.db.models import F
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.colors import HexColor
from reportlab.platypus import Image, Paragraph, Table, TableStyle
from reportlab.lib.units import mm, inch
from reportlab.lib import colors
import textwrap
from django.views.static import serve
from django.http import FileResponse
#from datetime import datetime
from num2words import num2words

@login_required(login_url="/employee/login/")
def rfp_pending_list(request):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        type = u.profile.type
        context['login_user_name'] = u.first_name + ' ' + u.last_name

        if type == 'Sourcing':
                if request.method == "GET":
                        rfp = RFP.objects.filter(opportunity_status='Open',enquiry_status='Approved',rfp_assign1__assign_to1=user).values('rfp_no','customer__name','customer__location','rfp_creation_details__creation_date','rfp_keyaccounts_details__key_accounts_manager__first_name','priority')    
                        context['rfp_list'] = rfp
                        return render(request,"Sourcing/Sourcing/pending_list.html",context)

@login_required(login_url="/employee/login/")
def rfp_pending_lineitems(request,rfp_no=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        type = u.profile.type
        context['login_user_name'] = u.first_name + ' ' + u.last_name

        if type == 'Sourcing':
                if request.method == "GET":
                        rfp_lineitems = RFPLineitem.objects.filter(rfp_no=rfp_no).order_by('creation_time')

                        rfp = RFP.objects.get(rfp_no=rfp_no)
                        context['rfp'] = rfp

                        
                        context['rfp_no'] = rfp_no
                        context['lineitems'] = rfp_lineitems
                        return render(request,"Sourcing/Sourcing/pending_lineitems.html",context)

@login_required(login_url="/employee/login/")
def lineitem_edit_tax(request,rfp_no=None,lineitem_id=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        type = u.profile.type
        context['login_user_name'] = u.first_name + ' ' + u.last_name

        if type == 'Sourcing':
                if request.method == "GET":
                        rfp_lineitem = RFPLineitem.objects.filter(lineitem_id=lineitem_id)[0]
                        context['rfp_no'] = rfp_no
                        context['lineitem'] = rfp_lineitem
                        return render(request,"Sourcing/Sourcing/lineitem_tax_edit.html",context)
            
                if request.method == "POST":
                        data = request.POST
                        rfp_lineitem = RFPLineitem.objects.get(lineitem_id=lineitem_id)
                        rfp_lineitem.hsn_code = data['hsn']
                        rfp_lineitem.gst = data['gst']
                        rfp_lineitem.save() 
                        return HttpResponseRedirect(reverse('rfp_pending_lineitems', args=[rfp_no]))

@login_required(login_url="/employee/login/")
def vendor_selection(request,rfp_no=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        type = u.profile.type
        context['login_user_name'] = u.first_name + ' ' + u.last_name

        if type == 'Sourcing':
                if request.method == "GET":
                        supplier_list = SupplierProfile.objects.all()
                        context['rfp_no'] = rfp_no

                        selected_supplier_list = Sourcing.objects.filter(rfp=rfp_no).values(
                                'id',
                                'supplier__name',
                                'supplier__location',
                                'supplier_contact_person__name',
                                'offer_reference',
                                'offer_date')
                        context['selected_supplier_list'] = selected_supplier_list

                        context['Supplier_list'] = supplier_list
                        state = StateList.objects.all()
                        context['StateList'] = state
                        return render(request,"Sourcing/Sourcing/vendor_selection.html",context)

@login_required(login_url="/employee/login/")
def new_vendor(request,rfp_no=None):
        context={}
        context['supplier'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        type = u.profile.type
        context['login_user_name'] = u.first_name + ' ' + u.last_name

        if type == 'Sourcing':
                if request.method == "POST":
                        user = User.objects.get(username=request.user)
                        data = request.POST
                        if data['country'].upper() == 'INDIA':
                                supplier_id = 'V9' + str(format(SupplierProfile.objects.count() + 1, '05d'))
                        else:
                                supplier_id = 'VI9' + str(format(SupplierProfile.objects.count() + 1, '05d'))

                        if data['PaymentTerm'] == '':
                                pt = 0
                        else:
                                pt=data['PaymentTerm']
        
                        if data['advance'] == '':
                                adv = 0
                        else:
                                adv = data['advance']
                        ven = SupplierProfile.objects.create(id=supplier_id,name=data['name'],location=data['location'],address=data['address'],city=data['city'],state=data['state'],pin=data['pin'],country=data['country'],office_email1=data['officeemail1'],office_email2=data['officeemail2'],office_phone1=data['officephone1'],office_phone2=data['officephone2'],gst_number=data['GSTNo'],advance_persentage=adv,payment_term=pt,inco_term=data['IncoTerm'],created_by=user)
                        if ven:
                                supplier_list = SupplierProfile.objects.all()
                                context['rfp_no'] = rfp_no
                                context['Supplier_list'] = supplier_list
                                state = StateList.objects.all()
                                context['StateList'] = state
                                return HttpResponseRedirect(reverse('vendor-selection', args=[rfp_no]))

@login_required(login_url="/employee/login/")
def vendor_contact_person_selection(request,rfp_no=None,vendor_id=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        type = u.profile.type
        context['login_user_name'] = u.first_name + ' ' + u.last_name
        context['rfp_no'] = rfp_no

        if type == 'Sourcing':
                if request.method == "GET":
                        ContactPerson = SupplierContactPerson.objects.filter(supplier_name__pk=vendor_id)
                        context['ContactPerson'] = ContactPerson
                        supplier = SupplierProfile.objects.get(id=vendor_id)
                        context['SupplierName'] = supplier.name
                        context['SupplierID'] = vendor_id
                        return render(request,"Sourcing/Sourcing/contact_person_selection.html",context)

                if request.method == "POST":
                        data = request.POST
                        supplier = SupplierProfile.objects.get(id=vendor_id)
                        contactperson_id =  vendor_id +'VP' + str(SupplierContactPerson.objects.count() + 1)
                        cp = SupplierContactPerson.objects.create(id=contactperson_id,name=data['name'],mobileNo1=data['phone1'],mobileNo2=data['phone2'],email1=data['email1'],email2=data['email2'],supplier_name=supplier,created_by=user)
                        ContactPerson = SupplierContactPerson.objects.filter(supplier_name__pk=vendor_id)
                        context['ContactPerson'] = ContactPerson
                        supplier = SupplierProfile.objects.get(id=vendor_id)
                        context['SupplierName'] = supplier.name
                        context['SupplierID'] = vendor_id
                        return render(request,"Sourcing/Sourcing/contact_person_selection.html",context)

@login_required(login_url="/employee/login/")
def offer_reference(request,rfp_no=None,vendor_id=None,contact_person_id=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        type = u.profile.type
        context['login_user_name'] = u.first_name + ' ' + u.last_name
        context['rfp_no'] = rfp_no
        context['supplier_id'] = vendor_id
        context['contact_person_id'] = contact_person_id

        if type == 'Sourcing':
                if request.method == "GET":
                        return render(request,"Sourcing/Sourcing/offer_reference.html",context)

                if request.method == "POST":
                        data = request.POST
                        sourcing_id = rfp_no + vendor_id
                        Sourcing.objects.create(id=sourcing_id,rfp=RFP.objects.get(rfp_no=rfp_no),supplier=SupplierProfile.objects.get(id=vendor_id),supplier_contact_person=SupplierContactPerson.objects.get(id=contact_person_id),offer_reference=data['reference'],offer_date=data['oDate'],created_by=user)
                        return HttpResponseRedirect(reverse('vendor-selection', args=[rfp_no]))

#Add Front Page Header
def Add_Header(pdf):
        pdf.drawInlineImage("static/image/aeprocurex.jpg",360,750,220,70)
        pdf.setFont('Helvetica-Bold', 13)
        pdf.drawString(10,763,"AEPROCUREX SOURCING PRIVATE LIMITED")

        pdf.setFont('Helvetica',9)
        #pdf.setFillColor(HexColor('#000000'))
        pdf.drawString(10,752,"Regd. Office: Shankarappa Complex, No.4")
        pdf.drawString(10,741,"Hosapalya Main Road, Opp. To Om Shakti Temple")
        pdf.drawString(10,730,"HSR Layout Extension,Bangalore - 560068")

        pdf.drawString(10,719,"Telephone: 080-43743314, +91 9964892600")
        pdf.drawString(10,708,"E-mail: sales.p@aeprocurex.com")
        pdf.drawString(10,697,"GST No. 29AAQCA2809L1Z6")
        pdf.drawString(10,686,"PAN No. - AAQCA2809L")
        pdf.drawString(10,675,"CIN No.-U74999KA2017PTC108349")

        pdf.setFont('Helvetica-Bold', 20)
        pdf.drawString(300,700,"REQUEST FOR QUOTATION")
        #pdf.setFillColor(yellow)
        pdf.rect(300,696,275,1, stroke=1, fill=1)

#Add Footer 
def Add_Footer(pdf):
        pdf.setFont('Helvetica', 10)
        pdf.drawString(250,15,'aeprocurex.com')
        pdf.drawString(520,15,'Page-No : ' + str(pdf.getPageNumber()))

#Add Request Information
def request_information(pdf,request_no, request_date,vendor_code):
        pdf.setFont('Helvetica', 9)
        pdf.drawString(300,680,"Request No :")
        pdf.drawString(300,667,"Request Date :")
        pdf.drawString(300,645,"Vendor Code :")

        pdf.setFont('Helvetica-Bold', 9)
        try:
                pdf.drawString(390,680,request_no)
        except:
                pdf.drawString(390,680,"")
        try:
                pdf.drawString(390,667,str(request_date.strftime('%d, %b %Y')))
        except:
                pdf.drawString(390,667,"")
        try:
                pdf.drawString(390,645,vendor_code)
        except:
                pdf.drawString(390,645,"")

#Add Request To
def Add_To(pdf,supplier_name,address,gst):
        pdf.setFont('Helvetica-Bold', 13)
        pdf.drawString(10,622,'REQUEST TO')
        pdf.rect(10,618,85,0.5, stroke=1, fill=1)
        #pdf.setFont('Helvetica-Bold', 13)

        pdf.setFont('Helvetica-Bold', 10)
        pdf.drawString(10,605,supplier_name)
        pdf.setFont('Helvetica', 9)
    
        wrapper = textwrap.TextWrapper(width=115) 
        word_list = wrapper.wrap(text=address)
        y = 590 
        for element in word_list:
                pdf.drawString(10,y,element)
                y = y - 13
        y = y -3
        try:
                pdf.drawString(10,y,'GST # :' + gst)
                y = y - 13
        except:
                pass
        y = y - 3
        return(y)

#ADD greeting
def Add_grt(pdf,y):
        pdf.setFont('Helvetica-Bold', 8)
        pdf.drawString(10,y,"Dear Sir / Madam,")
        y = y - 9
        pdf.drawString(30,y,"Please Provide the Price/s for following listed items as soon as possible")
        y = y - 9
        return(y)

#Add Table Header
def Add_Table_Header(pdf,y):
        pdf.rect(10,y,570,1, stroke=1, fill=1)
        pdf.setFillColor(HexColor('#E4E4E4'))
        pdf.rect(10,y-31,570,30, stroke=0, fill=1)
        #Colum headers
        pdf.setFillColor(HexColor('#000000'))
        pdf.setFont('Helvetica-Bold', 8)
        pdf.drawString(12,y-17,'SL #')
        pdf.drawString(60,y-17,'Material / Description / Specification')
        pdf.drawString(400,y-17,'Quantity')
        pdf.drawString(520,y-17,'UOM')
        y = y - 31
        pdf.rect(10,y,570,0.1, stroke=1, fill=1)
        y = y - 10
        return(y)

#Add Lineitem
def add_lineitem(pdf,y,i,request_number,product_title,description,model,brand,product_code,quantity,uom):
        
        pdf.setFont('Helvetica', 9)
        pdf.drawString(12,y,str(i))
        pdf.drawString(400,y,str(quantity))
        pdf.drawString(520,y,uom.upper())

        #Product title
        material_wrapper = textwrap.TextWrapper(width=45)
        title_word_list = material_wrapper.wrap(text=product_title)
    
        #Page Break
        if y < 50:
                y = add_new_page(pdf,request_number)
                y = Add_Table_Header(pdf,y)
                pdf.setFont('Helvetica', 9)

        for element in title_word_list:
                pdf.drawString(40,y,element)
                y = y - 11
        
        
                #Page Break
                if y < 50:
                        y = add_new_page(pdf,request_number)
                        y = Add_Table_Header(pdf,y)
                        pdf.setFont('Helvetica', 9)
        
        y = y + 10
        
        #Description
        description_word_list = material_wrapper.wrap(description)
        for element in description_word_list:
                pdf.drawString(40,y-10,element)
                y = y - 11
    
                #Page Break
                if y < 50:
                        y = add_new_page(pdf,request_number)
                        y = Add_Table_Header(pdf,y)
                        pdf.setFont('Helvetica', 9)
        y=y-3
        #Make
        if brand != '':
                try:
                        brand_word_list = material_wrapper.wrap('Make :' + brand)
                        for element in brand_word_list:
                
                                pdf.drawString(40,y-10,element)
                                y = y - 11

                                #Page Break
                                if y < 50:
                                        y = add_new_page(pdf,request_number)
                                        y = Add_Table_Header(pdf,y)
                                        pdf.setFont('Helvetica', 9)
                                
                        y=y-3
                except:
                        pass

        #Model
        if model != '':
                try:
                        model_word_list = material_wrapper.wrap('Model :' + model)
                        for element in model_word_list:
                
                                pdf.drawString(40,y-10,element)
                                y = y - 11

                                #Page Break
                                if y < 50:
                                        y = add_new_page(pdf,request_number)
                                        y = Add_Table_Header(pdf,y)
                                        pdf.setFont('Helvetica', 9)
                                
                        y=y-3
                except:
                        pass

        #Product Code
        if product_code != '':
                try:
                        product_code_word_list = material_wrapper.wrap('Product Code :' + product_code)
                        for element in product_code_word_list:
                
                                pdf.drawString(40,y-10,element)
                                y = y - 11

                                #Page Break
                                if y < 50:
                                        y = add_new_page(pdf,request_number)
                                        y = Add_Table_Header(pdf,y)
                                        pdf.setFont('Helvetica', 9)
                                
                        y=y-3
                except:
                        pass
        
        pdf.rect(10,y,568,0.1, stroke=1, fill=1)
        y = y - 10
        return(y)

#Add New Page
def add_new_page(pdf,request_number):
        pdf.showPage()
        pdf.drawInlineImage("static/image/aeprocurex.jpg",360,750,220,70)
        Add_Footer(pdf)
        pdf.setFont('Helvetica-Bold', 13)
        pdf.drawString(10,763,'Request NO : ' + request_number)
        pdf.line(10,748,580,748)
        return(740)

#add requester
def add_requester(pdf,y,request_number,requester_name,email,phone):
        #Page Break
        if y < 100:
                y = add_new_page(pdf,request_number)
                pdf.setFont('Helvetica', 9)
        pdf.setFont('Helvetica-Bold', 8)

        pdf.drawString(20,y,"Note : Kindly Mention Payment Terms, Inco Terms and Pack Size, MOQ, Lead Time, MRP, HSN Code and GST(%) of each item In Your Offer.")
        y = y - 15

        try:
                pdf.drawString(400,y,"Requester :")
                pdf.drawString(450,y,requester_name)
                y = y - 10
        except:
                y = y - 10
        
        try:
                pdf.drawString(400,y,"Email :")
                pdf.drawString(450,y,email)
                y = y - 10
        except:
                y = y - 10
        
        try:
                pdf.drawString(400,y,"Contact No:")
                pdf.drawString(450,y,phone)
                y = y - 10
        except:
                y = y - 20

        try:
                pdf.drawString(190,y,"[ THIS IS A SYSTEM GENERATED REQUEST ]")
                y = y - 10
        except:
                y = y - 10
        return(y)        

def Generate_RFQ(rfq_no):
        rfq = RFQ.objects.get(id = rfq_no)
        pdf = canvas.Canvas("media/rfq/" + rfq_no + ".pdf", pagesize=A4)
        pdf.setTitle(rfq_no + '.pdf')
        Add_Header(pdf)
        Add_Footer(pdf)
        request_information(
                pdf,
                rfq_no,
                rfq.date,
                rfq.sourcing.supplier.id
                )
        y = Add_To(
                pdf,
                rfq.sourcing.supplier.name,
                rfq.sourcing.supplier.address,
                rfq.sourcing.supplier.gst_number
                )
        y = Add_grt(
                pdf,
                y)
        y = Add_Table_Header(pdf,y)

        rfq_lineitem = RFQLineitem.objects.filter(rfq=rfq)

        i = 1
        for item in rfq_lineitem:
                product_title = item.rfp_lineitem.product_title
                description = item.rfp_lineitem.description
                model = item.rfp_lineitem.model
                brand = item.rfp_lineitem.brand
                product_code = item.rfp_lineitem.product_code

                quantity = item.rfp_lineitem.quantity
                uom = item.rfp_lineitem.uom

                y = add_lineitem(
                        pdf,
                        y,
                        i,
                        rfq_no,
                        product_title,
                        description,
                        model,
                        brand,
                        product_code,
                        quantity,
                        uom)
                i = i + 1

        add_requester(
                pdf,
                y,
                rfq_no,
                rfq.sourcing.rfp.rfp_assign1.assign_to1.first_name + ' ' + rfq.sourcing.rfp.rfp_assign1.assign_to1.last_name,
                rfq.sourcing.rfp.rfp_assign1.assign_to1.email,
                rfq.sourcing.rfp.rfp_assign1.assign_to1.profile.office_mobile
                )

        pdf.showPage()
        pdf.save()

@login_required(login_url="/employee/login/")
def rfq_product_selection(request,rfp_no=None,sourcing_id=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        type = u.profile.type
        context['login_user_name'] = u.first_name + ' ' + u.last_name
        context['rfp_no'] = rfp_no
        context['sourcing_id'] = sourcing_id

        if type == 'Sourcing':
                if request.method == "GET":
                        rfp = RFP.objects.get(rfp_no=rfp_no)
                        rfp_lineitems = RFPLineitem.objects.filter(rfp_no=rfp)
                        context['rfp_lineitems'] = rfp_lineitems
                        #print(rfp_lineitems)
                        return render(request,"Sourcing/Sourcing/RFQ/product_selection.html",context)

                if request.method == "POST":
                        data = request.POST
                        lineitem = data['lineitem']
                        lineitem_list = lineitem.split(",")
                        print(lineitem_list)   

                        sourcing = Sourcing.objects.get(id = sourcing_id)
                        #Delete existing RFQ
                        rfq_list = RFQ.objects.filter(sourcing = sourcing)
                        for rq in rfq_list:
                                rq.delete()

                        rfq_count = RFQ.objects.count()
                        rfq_id = sourcing.supplier.id + 'RFQ' +str(rfq_count + 1) 

                        print(rfq_id)
                        rfq = RFQ.objects.create(
                                id = rfq_id,
                                sourcing = sourcing
                        )

                        #RFQ Lineitem Creation
                        for item in lineitem_list:
                                if item != '':
                                        rfp_lineitem = RFPLineitem.objects.get(lineitem_id=item)
                                        RFQLineitem.objects.create(
                                                rfq = rfq,
                                                rfp_lineitem = rfp_lineitem 
                                        )
                        return JsonResponse(data)
                                                              
@login_required(login_url="/employee/login/")
def rfq_generate(request,rfp_no=None,sourcing_id=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        type = u.profile.type
        context['login_user_name'] = u.first_name + ' ' + u.last_name
        context['rfp_no'] = rfp_no
        context['sourcing_id'] = sourcing_id

        if type == 'Sourcing':
                if request.method == 'GET':
                        sourcing = Sourcing.objects.get(id = sourcing_id)
                        print(sourcing)

                        rfq = RFQ.objects.get(sourcing=sourcing)
                        Generate_RFQ(rfq.id)
                        context['rfq_id'] = rfq.id

                        return render(request,"Sourcing/Sourcing/RFQ/download_rfq.html",context)

@login_required(login_url="/employee/login/")
def single_price_request(request,rfp_no=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        context['login_user_name'] = u.first_name + ' ' + u.last_name
        type = u.profile.type

        if type == 'Sourcing':
                if request.method == "GET":
                        rfp_obj = RFP.objects.get(rfp_no=rfp_no)
                        if rfp_obj.single_vendor_approval == 'Approved':
                                context['error'] = 'The Enquiry is already approved for single vendor'
                                return render(request,"Sourcing/Sourcing/error.html",context)

                        return render(request,"Sourcing/Sourcing/single_price_request.html",context)

                if request.method == "POST":
                        data = request.POST
                        rfp_obj = RFP.objects.get(rfp_no=rfp_no)
                        rfp_obj.single_vendor_approval = 'Request'
                        rfp_obj.single_vendor_reason = data['reasons']
                        rfp_obj.save()

                        email_list = []
                        sales_team_email = Profile.objects.filter(type='Sales').values('user__email')
                        for email in sales_team_email:
                                email_list.append(email['user__email']) 

                        #Sending mail Notification
                        #email_receiver = rfp_obj.rfp_creation_details.created_by.email
                        lineitems = RFPLineitem.objects.filter(rfp_no=rfp_obj)
                        email_body = '<head>'\
                        '<style>'\
                        'table {'\
                        'width:100%;'\
                        '}'\
                        'table, th, td {'\
                        'border: 1px solid black;'\
                        'border-collapse: collapse;'\
                        '}'\
                        'th, td {'\
                        'padding: 15px;'\
                        'text-align: left;'\
                        '}'\
                        'table#t01 tr:nth-child(even) {'\
                        'background-color: #eee;'\
                        '}'\
                        'table#t01 tr:nth-child(odd) {'\
                        'background-color: #fff;'\
                        '}'\
                        'table#t01 th {'\
                        'background-color: #1E2DFF;'\
                        'color: white;'\
                        '}'\
                        '</style>'\
                        '</head>'\
                        '<body>'\
                        '<h1 style="text-align: center;"><span style="color: #0000ff;"><strong>AEPROCUREX ERP</strong></span></h1>'\
                        '<p><span style="color: #0000ff;"><strong> ' + request.user.first_name + ' ' + request.user.last_name + ' is asking your permission for single Vendor Sourcing'\
                        '</strong></span><span style="color: #0000ff;">'\
                        '<h4><strong>Reason :' + data['reasons'] + '</strong></h4>'\
                        '<p><span style="color: #0000ff;"><strong>RFP No : '+ rfp_no +'</strong></span></p>'\
                        '<p><span style="color: #0000ff;"><strong>Customer : '+ rfp_obj.customer.name +'</strong></span></p>'\
                        '<p><span style="color: #0000ff;"><strong>Requester : '+ rfp_obj.customer_contact_person.name +'</strong></span></p>'\
                        '<h2><span style="text-decoration: underline;"><span style="color: #000080; text-decoration: underline;">Enquiry Details :</span></span></h2>'\
                        '<table id="t01">'\
                        '<tr>'\
                        '<th align="Centre">Sl #</th>'\
                        '<th align="Centre">Product Title</th>'\
                        '<th align="Centre">Description</th>' \
                        '<th align="Centre">Quantity</th>'\
                        '<th align="Centre">UOM</th>'\
                        '</tr>'
                        i = 1
                        for items in lineitems:      
                                email_body = email_body + '<tr>'\
                                '<td>'+ str(i) +'</td>'\
                                '<td>'+ items.product_title +'</td>'\
                                '<td>'+ items.description +'</td>'\
                                '<td>'+ str(items.quantity) +'</td>'\
                                '<td>'+ items.uom +'</td>'\
                                '</tr>'
                                i = i + 1
                        email_body = email_body + '</table>'\
                        '<h2><span style="text-decoration: underline;"><span style="color: #000080; text-decoration: underline;">Sourcing Details :</span></span></h2>'
            
                        #Sourcing Details
                        sourcing_list = Sourcing.objects.filter(rfp=rfp_obj)
                        for sourcing in sourcing_list:
                                print(sourcing.supplier.name)
                                email_body = email_body + '<h3>Supplier Name : '+ sourcing.supplier.name +'</h3>'\
                                '<p>Location : '+ sourcing.supplier.location +'</p>'\
                                '<table id="t01">'\
                                '<tr>'\
                                '<th align="Centre">Sl #</th>'\
                                '<th align="Centre">Product Title</th>'\
                                '<th align="Centre">Description</th>' \
                                '<th align="Centre">MRP</th>'\
                                '<th align="Centre">Initial Price</th>'\
                                '<th align="Centre">Negotiate Price</th>'\
                                '<th align="Centre">Lead Time</th>'\
                                '</tr>'
                                i = 1
                                sourcing_lineitems = SourcingLineitem.objects.filter(sourcing=sourcing)
                                for item in sourcing_lineitems:      
                                        email_body = email_body + '<tr>'\
                                        '<td>'+ str(i) +'</td>'\
                                        '<td>'+ item.product_title +'</td>'\
                                        '<td>'+ item.description +'</td>'\
                                        '<td>'+ str(item.mrp) +'</td>'\
                                        '<td>'+ str(item.price1) +'</td>'\
                                        '<td>'+ str(item.price2) +'</td>'\
                                        '<td>'+ str(item.lead_time) +'</td>'\
                                        '</tr>'
                                        i = i + 1
                                email_body = email_body + '</table>'
                                email_body = email_body + '</body>'
                        msg = EmailMessage(subject=rfp_no, body=email_body, from_email = settings.DEFAULT_FROM_EMAIL,to = email_list, bcc = ['sales.p@aeprocurex.com','milan.kar@aeprocurex.com'])
                        msg.content_subtype = "html"  # Main content is now text/html
                        msg.send()
                        return HttpResponseRedirect(reverse('vendor-selection', args=[rfp_no]))

@login_required(login_url="/employee/login/")
def single_vendor_approval_list(request):
    context={}
    context['sourcing'] = 'active'
    user = User.objects.get(username=request.user)
    u = User.objects.get(username=request.user)
    context['login_user_name'] = u.first_name + ' ' + u.last_name
    type = u.profile.type

    if type == 'Sales':
        if request.method == "GET":
            rfp_list = RFP.objects.filter(single_vendor_approval='Request').values(
                'rfp_no',
                'customer__name',
                'customer__location',
                'customer_contact_person__name',
                'rfp_creation_details__creation_date',
                'rfp_assign1__assign_to1__first_name',
                'priority'
                )
            context['rfp_list'] = rfp_list
            print(rfp_list)
            return render(request,"Sales/Sourcing/single_vendor_approval_list.html",context)

@login_required(login_url="/employee/login/")
def single_vendor_approval_details(request,rfp_no=None):
    context={}
    context['sourcing'] = 'active'
    user = User.objects.get(username=request.user)
    u = User.objects.get(username=request.user)
    context['login_user_name'] = u.first_name + ' ' + u.last_name
    type = u.profile.type

    if type == 'Sales':
        if request.method == "GET":
            reason = RFP.objects.filter(rfp_no=rfp_no).values('single_vendor_reason')[0]['single_vendor_reason']
            context['reason'] = reason
            lineitem = RFPLineitem.objects.filter(rfp_no=RFP.objects.get(rfp_no=rfp_no))
            context['rfp_lineitems'] = lineitem
            return render(request,"Sales/Sourcing/single_vendor_approval_details.html",context)

@login_required(login_url="/employee/login/")
def single_vendor_approve(request,rfp_no=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        context['login_user_name'] = u.first_name + ' ' + u.last_name
        type = u.profile.type

        if type == 'Sales':
                if request.method == "POST":
                        rfp=RFP.objects.get(rfp_no=rfp_no)
                        rfp.single_vendor_approval = 'Approved'
                        rfp.save()
            
                        #Sending mail Notification
                        email_receiver = rfp.rfp_assign1.assign_to1.email
                        lineitems = RFPLineitem.objects.filter(rfp_no=rfp)
                        email_body = '<head>'\
                        '<style>'\
                        'table {'\
                        'width:100%;'\
                        '}'\
                        'table, th, td {'\
                        'border: 1px solid black;'\
                        'border-collapse: collapse;'\
                        '}'\
                        'th, td {'\
                        'padding: 15px;'\
                        'text-align: left;'\
                        '}'\
                        'table#t01 tr:nth-child(even) {'\
                        'background-color: #eee;'\
                        '}'\
                        'table#t01 tr:nth-child(odd) {'\
                        'background-color: #fff;'\
                        '}'\
                        'table#t01 th {'\
                        'background-color: #1E2DFF;'\
                        'color: white;'\
                        '}'\
                        '</style>'\
                        '</head>'\
                        '<body>'\
                        '<h1 style="text-align: center;"><span style="color: #0000ff;"><strong>AEPROCUREX ERP</strong></span></h1>'\
                        '<h4><strong>Hello, ' + rfp.rfp_assign1.assign_to1.first_name + ' ' + rfp.rfp_assign1.assign_to1.last_name + '</strong></h4>'\
                        '<p><span style="color: #0000ff;"><strong> ' + request.user.first_name + ' ' + request.user.last_name + ' has approved your request for single Vendor Sourcing'\
                        '</strong></span><span style="color: #0000ff;">'\
                        '<p><span style="color: #0000ff;"><strong>RFP No : '+ rfp_no +'</strong></span></p>'\
                        '<p><span style="color: #0000ff;"><strong>Customer : '+ rfp.customer.name +'</strong></span></p>'\
                        '<p><span style="color: #0000ff;"><strong>Requester : '+ rfp.customer_contact_person.name +'</strong></span></p>'\
                        '<h2><span style="text-decoration: underline;"><span style="color: #000080; text-decoration: underline;">Enquiry Details :</span></span></h2>'\
                        '<table id="t01">'\
                        '<tr>'\
                        '<th align="Centre">Sl #</th>'\
                        '<th align="Centre">Product Title</th>'\
                        '<th align="Centre">Description</th>' \
                        '<th align="Centre">Quantity</th>'\
                        '<th align="Centre">UOM</th>'\
                        '</tr>'
                        i = 1
                        for items in lineitems:      
                                email_body = email_body + '<tr>'\
                                '<td>'+ str(i) +'</td>'\
                                '<td>'+ items.product_title +'</td>'\
                                '<td>'+ items.description +'</td>'\
                                '<td>'+ str(items.quantity) +'</td>'\
                                '<td>'+ items.uom +'</td>'\
                                '</tr>'
                                i = i + 1
                        email_body = email_body + '</table>'\
                        '<h2><span style="text-decoration: underline;"><span style="color: #000080; text-decoration: underline;">Sourcing Details :</span></span></h2>'
            
                        #Sourcing Details
                        sourcing_list = Sourcing.objects.filter(rfp=rfp)
                        for sourcing in sourcing_list:
                                print(sourcing.supplier.name)
                                email_body = email_body + '<h3>Supplier Name : '+ sourcing.supplier.name +'</h3>'\
                                '<p>Location : '+ sourcing.supplier.location +'</p>'\
                                '<table id="t01">'\
                                '<tr>'\
                                '<th align="Centre">Sl #</th>'\
                                '<th align="Centre">Product Title</th>'\
                                '<th align="Centre">Description</th>' \
                                '<th align="Centre">MRP</th>'\
                                '<th align="Centre">Initial Price</th>'\
                                '<th align="Centre">Negotiate Price</th>'\
                                '<th align="Centre">Lead Time</th>'\
                                '</tr>'
                                i = 1
                                sourcing_lineitems = SourcingLineitem.objects.filter(sourcing=sourcing)
                                for item in sourcing_lineitems:      
                                        email_body = email_body + '<tr>'\
                                        '<td>'+ str(i) +'</td>'\
                                        '<td>'+ item.product_title +'</td>'\
                                        '<td>'+ item.description +'</td>'\
                                        '<td>'+ str(item.mrp) +'</td>'\
                                        '<td>'+ str(item.price1) +'</td>'\
                                        '<td>'+ str(item.price2) +'</td>'\
                                        '<td>'+ str(item.lead_time) +'</td>'\
                                        '</tr>'
                                        i = i + 1
                                email_body = email_body + '</table>'
                        email_body = email_body + '<p><span style="color: #ff0000;">Now you can mark sourcing complete with single vendor</span></p>'\
                        '</body>'
                        msg = EmailMessage(subject=rfp_no, body=email_body, from_email = settings.DEFAULT_FROM_EMAIL,to = [email_receiver], bcc = ['sales.p@aeprocurex.com','milan.kar@aeprocurex.com'])
                        msg.content_subtype = "html"  # Main content is now text/html
                        msg.send()
                        return HttpResponseRedirect(reverse('single-vendor-approval-list'))

@login_required(login_url="/employee/login/")
def single_vendor_reject(request,rfp_no=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        context['login_user_name'] = u.first_name + ' ' + u.last_name
        type = u.profile.type

        if type == 'Sales':
                if request.method == "POST":
                        rfp=RFP.objects.get(rfp_no=rfp_no)
                        rfp.single_vendor_approval = 'Rejected'
                        rfp.save()

                        #Sending mail Notification
                        email_receiver = rfp.rfp_assign1.assign_to1.email
                        lineitems = RFPLineitem.objects.filter(rfp_no=rfp)
                        email_body = '<head>'\
                        '<style>'\
                        'table {'\
                        'width:100%;'\
                        '}'\
                        'table, th, td {'\
                        'border: 1px solid black;'\
                        'border-collapse: collapse;'\
                        '}'\
                        'th, td {'\
                        'padding: 15px;'\
                        'text-align: left;'\
                        '}'\
                        'table#t01 tr:nth-child(even) {'\
                        'background-color: #eee;'\
                        '}'\
                        'table#t01 tr:nth-child(odd) {'\
                        'background-color: #fff;'\
                        '}'\
                        'table#t01 th {'\
                        'background-color: #1E2DFF;'\
                        'color: white;'\
                        '}'\
                        '</style>'\
                        '</head>'\
                        '<body>'\
                        '<h1 style="text-align: center;"><span style="color: #0000ff;"><strong>AEPROCUREX ERP</strong></span></h1>'\
                        '<h4><strong>Hello, ' + rfp.rfp_assign1.assign_to1.first_name + ' ' + rfp.rfp_assign1.assign_to1.last_name + '</strong></h4>'\
                        '<p><span style="color: #0000ff;"><strong> ' + request.user.first_name + ' ' + request.user.last_name + ' has rejected your request for single Vendor Sourcing'\
                        '</strong></span><span style="color: #0000ff;">'\
                        '<p><span style="color: #0000ff;"><strong>RFP No : '+ rfp_no +'</strong></span></p>'\
                        '<p><span style="color: #0000ff;"><strong>Customer : '+ rfp.customer.name +'</strong></span></p>'\
                        '<p><span style="color: #0000ff;"><strong>Requester : '+ rfp.customer_contact_person.name +'</strong></span></p>'\
                        '<h2><span style="text-decoration: underline;"><span style="color: #000080; text-decoration: underline;">Enquiry Details :</span></span></h2>'\
                        '<table id="t01">'\
                        '<tr>'\
                        '<th align="Centre">Sl #</th>'\
                        '<th align="Centre">Product Title</th>'\
                        '<th align="Centre">Description</th>' \
                        '<th align="Centre">Quantity</th>'\
                        '<th align="Centre">UOM</th>'\
                        '</tr>'
                        i = 1
                        for items in lineitems:      
                                email_body = email_body + '<tr>'\
                                '<td>'+ str(i) +'</td>'\
                                '<td>'+ items.product_title +'</td>'\
                                '<td>'+ items.description +'</td>'\
                                '<td>'+ str(items.quantity) +'</td>'\
                                '<td>'+ items.uom +'</td>'\
                                '</tr>'
                                i = i + 1
                        email_body = email_body + '</table>'\
                        '<h2><span style="text-decoration: underline;"><span style="color: #000080; text-decoration: underline;">Sourcing Details :</span></span></h2>'
            
                        #Sourcing Details
                        sourcing_list = Sourcing.objects.filter(rfp=rfp)
                        for sourcing in sourcing_list:
                                print(sourcing.supplier.name)
                                email_body = email_body + '<h3>Supplier Name : '+ sourcing.supplier.name +'</h3>'\
                                '<p>Location : '+ sourcing.supplier.location +'</p>'\
                                '<table id="t01">'\
                                '<tr>'\
                                '<th align="Centre">Sl #</th>'\
                                '<th align="Centre">Product Title</th>'\
                                '<th align="Centre">Description</th>' \
                                '<th align="Centre">MRP</th>'\
                                '<th align="Centre">Initial Price</th>'\
                                '<th align="Centre">Negotiate Price</th>'\
                                '<th align="Centre">Lead Time</th>'\
                                '</tr>'
                                i = 1
                                sourcing_lineitems = SourcingLineitem.objects.filter(sourcing=sourcing)
                                for item in sourcing_lineitems:      
                                        email_body = email_body + '<tr>'\
                                        '<td>'+ str(i) +'</td>'\
                                        '<td>'+ item.product_title +'</td>'\
                                        '<td>'+ item.description +'</td>'\
                                        '<td>'+ str(item.mrp) +'</td>'\
                                        '<td>'+ str(item.price1) +'</td>'\
                                        '<td>'+ str(item.price2) +'</td>'\
                                        '<td>'+ str(item.lead_time) +'</td>'\
                                        '</tr>'
                                        i = i + 1
                                email_body = email_body + '</table>'
                        email_body = email_body + '<p><span style="color: #ff0000;">Please Search Some other Vendors for this enquiry</span></p>'\
                        '</body>'
                        msg = EmailMessage(subject=rfp_no, body=email_body, from_email = settings.DEFAULT_FROM_EMAIL,to = [email_receiver], bcc = ['sales.p@aeprocurex.com','milan.kar@aeprocurex.com'])
                        msg.content_subtype = "html"  # Main content is now text/html
                        msg.send()

                        return HttpResponseRedirect(reverse('single-vendor-approval-list'))

@login_required(login_url="/employee/login/")
def single_vendor_history(request):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        context['login_user_name'] = u.first_name + ' ' + u.last_name
        type = u.profile.type

        if type == 'Sales':
                if request.method == "GET":
                        rfp_list = RFP.objects.exclude(single_vendor_approval = 'No').values(
                                'single_vendor_approval',
                                'rfp_no',
                                'customer__name',
                                'customer__location',
                                'customer_contact_person__name',
                                'rfp_creation_details__creation_date',
                                'rfp_assign1__assign_to1__first_name',
                                'single_vendor_reason',
                                'priority'
                                )
                        context['rfp_list'] = rfp_list
                        print(rfp_list)
                        return render(request,"Sales/Sourcing/single_vendor_history.html",context)  

@login_required(login_url="/employee/login/")
def vendor_quotation_edit(request,rfp_no=None,sourcing_id=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        context['login_user_name'] = u.first_name + ' ' + u.last_name
        type = u.profile.type

        if type == 'Sourcing':
                if request.method == "GET":
                        context['rfp_no'] = rfp_no
                        rfp_lineitems = RFPLineitem.objects.filter(rfp_no=rfp_no)
                        context['lineitems'] = rfp_lineitems

                        sourcing_lineitems = SourcingLineitem.objects.filter(sourcing__id=sourcing_id)
                        context['sourcing_lineitems'] = sourcing_lineitems

                        context['supplier_name'] = Sourcing.objects.filter(id=sourcing_id).values('supplier__name')[0]['supplier__name']
                        return render(request,"Sourcing/Sourcing/supplier_quotation_details.html",context)

@login_required(login_url="/employee/login/")
def vendor_quotation_price_upload(request,rfp_no=None,sourcing_id=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        context['login_user_name'] = u.first_name + ' ' + u.last_name
        type = u.profile.type

        if type == 'Sourcing':
                if request.method == "GET":
                        context['rfp_no'] = rfp_no
                        context['sourcing_id'] = sourcing_id
                        try:
                                wb = openpyxl.load_workbook(filename = 'media/xl_template/SourcingTemplate.xlsx')
                                worksheet = wb['ProductDetails']
                                worksheet['B1'] = sourcing_id
                                rfp_lineitems = RFPLineitem.objects.filter(rfp_no=RFP.objects.get(rfp_no=rfp_no)).order_by('creation_time')
                                i = 1
                                for item in rfp_lineitems:
                                        worksheet.append([
                                                i,
                                                item.lineitem_id,
                                                item.product_title,
                                                item.description,
                                                item.model,
                                                item.brand,
                                                item.product_code])
                                        i = i + 1
                                wb.save('media/xl_template/' + sourcing_id+'.xlsx')
                                return render(request,"Sourcing/Sourcing/upload_sourcing.html",context)
                        except:
                                pass
                
                if request.method == 'POST':
                        data_file = request.FILES['sourcing_file']
                        wb = openpyxl.load_workbook(data_file)
                        worksheet = wb['ProductDetails']
                        i = 1
                        for row in worksheet.iter_rows():
                                if i > 2 :
                                        if row[1].value != '' and row[12].value != '' and row[13].value != '' and row[7].value != '':
                                                sourcing_lineitem_id = sourcing_id + str(random.randint(100000,9999999))
                                                try :
                                                        rfp_lineitem_obj = RFPLineitem.objects.get(lineitem_id=row[1].value)
                                                        SourcingLineitem.objects.create(
                                                                id=sourcing_lineitem_id,
                                                                sourcing=Sourcing.objects.get(id=sourcing_id),
                                                                rfp_lineitem=rfp_lineitem_obj,
                                                                product_title=row[2].value,
                                                                description=row[3].value,
                                                                model=row[4].value,
                                                                brand=row[5].value,
                                                                product_code=row[6].value,
                                                                pack_size = row[7].value,
                                                                moq = row[8].value,
                                                                lead_time = row[9].value,
                                                                price_validity = row[10].value,
                                                                expected_freight = row[14].value,
                                                                mrp = row[11].value,
                                                                price1 = row[12].value,
                                                                price2 = row[13].value,
                                                                creation_time = rfp_lineitem_obj.creation_time )
                                                        print(rfp_lineitem_obj.creation_time)
                                                except:
                                                        pass
                                i = i + 1
                        return HttpResponseRedirect(reverse('vendor-quotation-edit', args=[rfp_no,sourcing_id]))

@login_required(login_url="/employee/login/")
def vendor_quotation_view(request,rfp_no=None,sourcing_id=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        context['login_user_name'] = u.first_name + ' ' + u.last_name
        type = u.profile.type

        if type == 'Sourcing':
                if request.method == "GET":
                        context['supplier_name'] = Sourcing.objects.filter(id=sourcing_id).values('supplier__name')[0]['supplier__name']
                        sourcing_lineitems = SourcingLineitem.objects.filter(sourcing__id=sourcing_id)
                        context['sourcing_lineitems'] = sourcing_lineitems
                        return render(request,"Sourcing/Sourcing/supplier_quotation_view.html",context)

@login_required(login_url="/employee/login/")
def vendor_quotation_delete(request,rfp_no=None,sourcing_id=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        context['login_user_name'] = u.first_name + ' ' + u.last_name
        type = u.profile.type

        if type == 'Sourcing':
                if request.method == "GET":
                        context['sourcing_details'] = Sourcing.objects.filter(id=sourcing_id).values('supplier__name','supplier_contact_person__name','offer_reference','offer_date')[0]
                        return render(request,"Sourcing/Sourcing/supplier_quotation_delete.html",context)

                if request.method == "POST":
                        sourcing_obj = Sourcing.objects.get(id=sourcing_id)
                        sourcing_obj.delete()
                        return HttpResponseRedirect(reverse('vendor-selection', args=[rfp_no]))

@login_required(login_url="/employee/login/")
def vendor_quotation_price_add(request,rfp_no=None,sourcing_id=None,lineitem_id=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        context['login_user_name'] = u.first_name + ' ' + u.last_name
        type = u.profile.type

        if type == 'Sourcing':
                if request.method == "GET":
                        rfp_lineitem = RFPLineitem.objects.filter(lineitem_id=lineitem_id)[0]
                        context['lineitem'] = rfp_lineitem
                        context['supplier_name'] = Sourcing.objects.filter(id=sourcing_id).values('supplier__name')[0]['supplier__name']
                        return render(request,"Sourcing/Sourcing/supplier_price_add.html",context)

                if request.method == "POST":
                        rfp_lineitem_obj = RFPLineitem.objects.get(lineitem_id=lineitem_id)
                        price_detail = request.POST
                        sourcing_lineitem_id = sourcing_id + str(random.randint(100000,9999999))
            
                        if price_detail['freight'] =='':
                                freight=0
                        else:
                                freight = float(price_detail['freight'])

                        if price_detail['mrp'] == '':
                                mrp=0
                        else:
                                mrp=price_detail['mrp']

                        if price_detail['price1'] == '':
                                price1 = 0
                        else:
                                price1 = price_detail['price1']

                        if price_detail['price2'] == '':
                                price2 = 0
                        else:
                                price2 = price_detail['price2']

                        SourcingLineitem.objects.create(
                                id=sourcing_lineitem_id,
                                sourcing=Sourcing.objects.get(id=sourcing_id),
                                rfp_lineitem=RFPLineitem.objects.get(lineitem_id=lineitem_id),
                                product_title=price_detail['product_title'],
                                description=price_detail['description'],
                                model=price_detail['model'],
                                brand=price_detail['brand'],
                                product_code=price_detail['product_code'],
                                pack_size = price_detail['pack_size'],
                                moq = price_detail['moq'],
                                lead_time = price_detail['lead_time'],
                                price_validity = price_detail['price_validity'],
                                expected_freight = freight,
                                mrp = mrp,
                                price1 = price1,
                                price2 = price2,
                                creation_time = rfp_lineitem_obj.creation_time )
                        #print(rfp_lineitem_obj.creation_time)
                        return HttpResponseRedirect(reverse('vendor-quotation-edit', args=[rfp_no,sourcing_id]))

@login_required(login_url="/employee/login/")
def vendor_quotation_price_edit(request,rfp_no=None,sourcing_id=None,price_id=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        context['login_user_name'] = u.first_name + ' ' + u.last_name
        type = u.profile.type

        if type == 'Sourcing':

                if request.method == "GET":
                        price = SourcingLineitem.objects.filter(id=price_id)[0]
                        context['price'] = price
                        return render(request,"Sourcing/Sourcing/supplier_price_edit.html",context)

                if request.method == "POST":
                        price_detail = request.POST
            
                        price_object = SourcingLineitem.objects.get(id=price_id)
            
                        if price_detail['freight'] =='':
                                freight=0
                        else:
                                freight = float(price_detail['freight'])

                        if price_detail['mrp'] == '':
                                mrp=0
                        else:
                                mrp=price_detail['mrp']

                        if price_detail['price1'] == '':
                                price1 = 0
                        else:
                                price1 = price_detail['price1']

                        if price_detail['price2'] == '':
                                price2 = 0
                        else:
                                price2 = price_detail['price2']

                        price_object.product_title=price_detail['product_title']
                        price_object.description=price_detail['description']
                        price_object.model=price_detail['model']
                        price_object.brand=price_detail['brand']
                        price_object.product_code=price_detail['product_code']
                        price_object.pack_size = price_detail['pack_size']
                        price_object.moq = price_detail['moq']
                        price_object.lead_time = price_detail['lead_time']
                        price_object.price_validity = price_detail['price_validity']
                        price_object.expected_freight = freight
                        price_object.mrp = mrp
                        price_object.price1 = price1
                        price_object.price2 = price2
            
                        price_object.save()
                        return HttpResponseRedirect(reverse('vendor-quotation-edit', args=[rfp_no,sourcing_id]))

@login_required(login_url="/employee/login/")
def vendor_quotation_price_delete(request,rfp_no=None,sourcing_id=None,price_id=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        type = u.profile.type
        context['login_user_name'] = u.first_name + ' ' + u.last_name

        if type == 'Sourcing':
                if request.method == "GET":
                        price = SourcingLineitem.objects.filter(id=price_id)[0]
                        context['price'] = price
                        return render(request,"Sourcing/Sourcing/supplier_price_delete.html",context)

                if request.method == "POST":
                        price_object = SourcingLineitem.objects.get(id=price_id)
                        price_object.delete()
                        return HttpResponseRedirect(reverse('vendor-quotation-edit', args=[rfp_no,sourcing_id]))

@login_required(login_url="/employee/login/")
def round2(request,rfp_no=None,sourcing_id=None,price_id=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        type = u.profile.type
        context['login_user_name'] = u.first_name + ' ' + u.last_name

        if type == 'Sourcing':
                if request.method == "GET":
                        price = SourcingLineitem.objects.filter(id=price_id)[0]
                        context['price'] = price
                        return render(request,"Sourcing/Sourcing/round2.html",context)

                if request.method=="POST":
                        data = request.POST
                        sourcing_obj = SourcingLineitem.objects.get(id=price_id)
                        sourcing_obj.price2 = data['round2']
                        sourcing_obj.save()
                        return HttpResponseRedirect(reverse('vendor-quotation-edit', args=[rfp_no,sourcing_id]))

@login_required(login_url="/employee/login/")
def sourcing_completed(request,rfp_no=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        type = u.profile.type
        context['login_user_name'] = u.first_name + ' ' + u.last_name

        if type == 'Sourcing':
                if request.method == "POST":

                        single_vendor_approval = RFP.objects.get(rfp_no=rfp_no)
                        sva = 'No'
                        if single_vendor_approval.single_vendor_approval == 'Approved':
                                sva = 'Yes'

                        #Temporary line
                        sva = 'Yes'
                        #Temporary Line Completed

                        #Checking GST & HSN
                        
                        rfp_lineitem = RFPLineitem.objects.filter(rfp_no=RFP.objects.get(rfp_no=rfp_no))
                        for item in rfp_lineitem:

                                if item.gst < 1  or item.hsn_code == '':
                                        context['error'] = "HSN or GST % Missing"
                                        return render(request,"Sourcing/Sourcing/error.html",context)
                        #end of Checking GST / HSN

                        supplier_quotation_count = Sourcing.objects.filter(rfp__rfp_no=rfp_no).count()
                        if supplier_quotation_count < 2: 
                                if sva == 'No':
                                        context['error'] = "You have to collect atleast two Quotation from vendor"
                                        return render(request,"Sourcing/Sourcing/error.html",context)

                        sourcing_count = SourcingLineitem.objects.filter(rfp_lineitem__rfp_no__rfp_no=rfp_no).count()
                        if sourcing_count == 0:
                                context['error'] = "No Price found"
                                return render(request,"Sourcing/Sourcing/error.html",context)

                        ##Price CHeck by lineitem
                        rfp_lineitems = RFPLineitem.objects.filter(rfp_no=rfp_no)
                        sourcing_lineitems = SourcingLineitem.objects.filter(sourcing__rfp__rfp_no=rfp_no).order_by('price2')

                        if sva == 'No':
                                for lineitem in rfp_lineitems:
                                        counter = 0
                                        for item in sourcing_lineitems:
                                                if lineitem.lineitem_id == item.rfp_lineitem.lineitem_id:
                                                        counter = counter + 1
                                                if counter < 2:
                                                        context['error'] = "Some lineitems found with less than two prices. Please Add atleast two price for each lineitem"
                                                        return render(request,"Sourcing/Sourcing/error.html",context)
                        ##Price check by lineitem

                        round2_count = SourcingLineitem.objects.filter(rfp_lineitem__rfp_no__rfp_no=rfp_no,price2=None).count()
                        if round2_count == 0:
                                data = request.POST    
                                sourcing_completed_details = RFPSourcingDetail.objects.create(id=rfp_no+str(random.randint(1000,99999)),sourcing_completed_by = request.user)
                                rfp = RFP.objects.get(rfp_no=rfp_no)
                                rfp.enquiry_status = 'Sourcing_Completed'
                                rfp.rfp_sourcing_detail = sourcing_completed_details
                                rfp.save()

                                email_list = []
                                sales_team_email = Profile.objects.filter(type='Sales').values('user__email')
                                for email in sales_team_email:
                                        email_list.append(email['user__email'])
                                print(email_list)
                                #Sending mail Notification
                                #email_receiver = rfp.rfp_creation_details.created_by.email
                                lineitems = RFPLineitem.objects.filter(rfp_no=rfp)
                                email_body = '<head>'\
                                '<style>'\
                                'table {'\
                                'width:100%;'\
                                '}'\
                                'table, th, td {'\
                                'border: 1px solid black;'\
                                'border-collapse: collapse;'\
                                '}'\
                                'th, td {'\
                                'padding: 15px;'\
                                'text-align: left;'\
                                '}'\
                                'table#t01 tr:nth-child(even) {'\
                                'background-color: #eee;'\
                                '}'\
                                'table#t01 tr:nth-child(odd) {'\
                                'background-color: #fff;'\
                                '}'\
                                'table#t01 th {'\
                                'background-color: #1E2DFF;'\
                                'color: white;'\
                                '}'\
                                '</style>'\
                                '</head>'\
                                '<body>'\
                                '<h1 style="text-align: center;"><span style="color: #0000ff;"><strong>AEPROCUREX ERP</strong></span></h1>'\
                                '<p><span style="color: #0000ff;"><strong> ' + request.user.first_name + ' ' + request.user.last_name + ' has marked one RFP as Sourcing Completed'\
                                '</strong></span><span style="color: #0000ff;">'\
                                '<p><span style="color: #0000ff;"><strong>RFP No : '+ rfp_no +'</strong></span></p>'\
                                '<p><span style="color: #0000ff;"><strong>Customer : '+ rfp.customer.name + ' - ' + rfp.customer.location +'</strong></span></p>'\
                                '<p><span style="color: #0000ff;"><strong>Requester : '+ rfp.customer_contact_person.name +'</strong></span></p>'\
                                '<h2><span style="text-decoration: underline;"><span style="color: #000080; text-decoration: underline;">Enquiry Details :</span></span></h2>'\
                                '<table id="t01">'\
                                '<tr>'\
                                '<th align="Centre">Sl #</th>'\
                                '<th align="Centre">Product Title</th>'\
                                '<th align="Centre">Description</th>' \
                                '<th align="Centre">Quantity</th>'\
                                '<th align="Centre">UOM</th>'\
                                '</tr>'
                                i = 1
                                for items in lineitems:      
                                        email_body = email_body + '<tr>'\
                                        '<td>'+ str(i) +'</td>'\
                                        '<td>'+ items.product_title +'</td>'\
                                        '<td>'+ items.description +'</td>'\
                                        '<td>'+ str(items.quantity) +'</td>'\
                                        '<td>'+ items.uom +'</td>'\
                                        '</tr>'
                                        i = i + 1
                                email_body = email_body + '</table>'\
                                '<h2><span style="text-decoration: underline;"><span style="color: #000080; text-decoration: underline;">Sourcing Details :</span></span></h2>'
            
                                #Sourcing Details
                                sourcing_list = Sourcing.objects.filter(rfp=rfp)
                                for sourcing in sourcing_list:
                                        print(sourcing.supplier.name)
                                        email_body = email_body + '<h3>Supplier Name : '+ sourcing.supplier.name +'</h3>'\
                                        '<p>Location : '+ sourcing.supplier.location +'</p>'\
                                        '<table id="t01">'\
                                        '<tr>'\
                                        '<th align="Centre">Sl #</th>'\
                                        '<th align="Centre">Product Title</th>'\
                                        '<th align="Centre">Description</th>' \
                                        '<th align="Centre">MRP</th>'\
                                        '<th align="Centre">Initial Price</th>'\
                                        '<th align="Centre">Negotiate Price</th>'\
                                        '<th align="Centre">Lead Time</th>'\
                                        '</tr>'
                                        i = 1
                                        sourcing_lineitems = SourcingLineitem.objects.filter(sourcing=sourcing)
                                        for item in sourcing_lineitems:      
                                                email_body = email_body + '<tr>'\
                                                '<td>'+ str(i) +'</td>'\
                                                '<td>'+ item.product_title +'</td>'\
                                                '<td>'+ item.description +'</td>'\
                                                '<td>'+ str(item.mrp) +'</td>'\
                                                '<td>'+ str(item.price1) +'</td>'\
                                                '<td>'+ str(item.price2) +'</td>'\
                                                '<td>'+ str(item.lead_time) +'</td>'\
                                                '</tr>'
                                                i = i + 1
                                        email_body = email_body + '</table>'\
                    
                                email_body = email_body + '</body>'
                                msg = EmailMessage(subject=rfp_no, body=email_body, from_email = settings.DEFAULT_FROM_EMAIL,to = email_list, bcc = ['sales.p@aeprocurex.com','milan.kar@aeprocurex.com'])
                                msg.content_subtype = "html"  # Main content is now text/html
                                msg.send()
                                context['message'] = 'RFP No ' + rfp_no + ' has been maked as sourcing completed successfully'
                                return render(request,"Sourcing/Sourcing/success.html",context)
                        else:
                                context['error'] = "Round 2 Price Not found"
                                return render(request,"Sourcing/Sourcing/error.html",context)
            
##Sourcing History
@login_required(login_url="/employee/login/")
def sourcing_item_wise(request,rfp_no=None):
        context={}
        context['sourcing'] = 'active'
        user = User.objects.get(username=request.user)
        u = User.objects.get(username=request.user)
        type = u.profile.type
        context['login_user_name'] = u.first_name + ' ' + u.last_name

        if request.method == 'GET':
                item_list = SourcingLineitem.objects.all()

                context['item_list'] = item_list

                if type == 'Sourcing':
                        return render(request,"Sourcing/Sourcing/History/item_list.html",context)

                if type == 'Sales':
                        return render(request,"Sales/Sourcing/History/item_list.html",context)